"""
Парсер входного Excel-файла (input.xlsx).
Возвращает все объекты, необходимые для планирования.
"""
from __future__ import annotations
import re
from typing import List, Tuple, Dict, Optional
import openpyxl

from models import (
    RawWire, InsulatedCore, CableStock, CableOrder,
    DrumType, CoreDrumCapacity, CableDrumCapacity, ProcessParams,
)


# ─── Вспомогательные функции ─────────────────────────────────────────

def _str(val) -> str:
    """Безопасное преобразование значения ячейки в строку."""
    if val is None:
        return ''
    return str(val).strip()


def _float(val, default: float = 0.0) -> float:
    """Безопасное преобразование в float."""
    if val is None:
        return default
    try:
        return float(str(val).replace(',', '.').replace(' ', ''))
    except (ValueError, TypeError):
        return default


def _parse_journal(raw: str) -> List[float]:
    """
    Разбирает строку кабельного журнала в список длин.
    '2000, 1500, 300' → [2000.0, 1500.0, 300.0]
    Пустая строка → []
    """
    if not raw or not _str(raw):
        return []
    tokens = re.split(r'[,;\s]+', _str(raw))
    result = []
    for t in tokens:
        t = t.strip()
        if t:
            try:
                result.append(float(t))
            except ValueError:
                pass  # пропускаем нечисловые токены
    return result


def _cross_section_str(val) -> str:
    """
    Нормализует сечение жилы:
      2.5  → '2,5'
      25.0 → '25'
      '2,5' → '2,5'
      '2.5' → '2,5'
    """
    if val is None:
        return ''
    s = str(val).strip()
    if not s:
        return ''
    # Если значение хранится как float (из Excel), нормализуем через float
    try:
        f = float(s.replace(',', '.'))
        if f == int(f):
            return str(int(f))
        return (f'{f:.4g}').replace('.', ',')
    except (ValueError, TypeError):
        # Уже строка типа '2,5' — просто вернём
        return s


def _drum_type_name(header: str) -> str:
    """
    Очищает заголовок столбца барабана от суффикса ', м'.
    'Б-630, м' → 'Б-630'
    '№12, м'  → '№12'
    """
    return re.sub(r',?\s*м\s*$', '', _str(header)).strip()


def _normalize_fr(val: str) -> str:
    """Нормализует признак огнестойкости: 'FR' → 'FR', '—'/'' → ''."""
    s = val.strip().upper()
    return 'FR' if s == 'FR' else ''


# ─── Парсинг каждого листа ───────────────────────────────────────────

def _parse_orders(ws) -> List[CableOrder]:
    """
    Лист '1. Заказы'.
    Строка 1 — заголовок листа, строка 2 — шапка, строки 3+ — данные.
    Колонки: A=Марка, B=Длина, C=Журнал
    """
    orders = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        mark = _str(row[0])
        if not mark:
            continue
        length = _float(row[1])
        journal_raw = _str(row[2]) if len(row) > 2 else ''
        journal = _parse_journal(journal_raw)

        # Цвета и атрибуты жил заполним позже (merge с листом Состав кабелей)
        orders.append(CableOrder(
            mark=mark,
            total_length=length,
            journal=journal,
            colors=[],
            cross_section='',
        ))
    return orders


def _parse_composition(ws) -> Dict[str, Tuple]:
    """
    Лист '2. Состав кабелей'.
    Колонки (строка 2 — шапка):
      A=Марка кабеля, B=Сечение жил, C=Индекс тпж, D=Огнестойкость,
      E=Материал изоляции, F..J=Жила 1..5, K=Примечание
    Возвращает:
      {марка: (cross_section, wire_type, fire_resistant, insulation_material, [цвета])}
    """
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    color_col_indices = []
    if header_row:
        for i in range(len(header_row)):
            h = _str(header_row[i])
            if h.lower().startswith('жила'):
                color_col_indices.append(i)

    composition: Dict[str, Tuple] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        mark = _str(row[0])
        if not mark:
            continue
        cross_section       = _cross_section_str(row[1]) if len(row) > 1 else ''
        wire_type           = _str(row[2]) if len(row) > 2 else ''
        fire_resistant      = _normalize_fr(_str(row[3]) if len(row) > 3 else '')
        insulation_material = _str(row[4]) if len(row) > 4 else ''

        colors = []
        for i in color_col_indices:
            if i < len(row):
                c = _str(row[i])
                if c:
                    colors.append(c)

        composition[mark] = (cross_section, wire_type, fire_resistant, insulation_material, colors)
    return composition


def _parse_pf(ws) -> Tuple[List[RawWire], List[InsulatedCore], List[CableStock]]:
    """
    Лист '3. П-Ф (склад)'.
    Шапка (строка 2):
      A=№  B=Тип  C=Сечение, мм²  D=Индекс  E=Огнестойкость
      F=Материал изол.  G=Цвет / Марка кабеля  H=Длина, м  I=Примечание (ID)

    Типы строк: 'ТПЖ', 'Изолированная', 'Кабель'
    """
    raw_wires: List[RawWire] = []
    insulated: List[InsulatedCore] = []
    cable_stock: List[CableStock] = []

    counters = {'ТПЖ': 0, 'Изолированная': 0, 'Кабель': 0}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        pf_type             = _str(row[1]) if len(row) > 1 else ''
        cross_section       = _cross_section_str(row[2]) if len(row) > 2 else ''
        wire_type           = _str(row[3]) if len(row) > 3 else ''
        fire_resistant      = _normalize_fr(_str(row[4]) if len(row) > 4 else '')
        insulation_material = _str(row[5]) if len(row) > 5 else ''
        name_or_mark        = _str(row[6]) if len(row) > 6 else ''
        length              = _float(row[7]) if len(row) > 7 else 0.0
        note                = _str(row[8]) if len(row) > 8 else ''

        if pf_type not in ('ТПЖ', 'Изолированная', 'Кабель'):
            continue
        if length <= 0:
            continue

        counters[pf_type] += 1
        item_id = note if note else f'{pf_type}-{counters[pf_type]:03d}'

        if pf_type == 'ТПЖ':
            wire_name = f'ТПЖ {cross_section}{wire_type}'
            raw_wires.append(RawWire(
                id=item_id,
                name=wire_name,
                cross_section=cross_section,
                wire_type=wire_type,
                length=length,
            ))

        elif pf_type == 'Изолированная':
            color = name_or_mark
            ins_name = f'{color} {cross_section}{wire_type}'
            if fire_resistant:
                ins_name += f' {fire_resistant}'
            ins_name += f' {insulation_material}'
            insulated.append(InsulatedCore(
                id=item_id,
                name=ins_name,
                color=color,
                cross_section=cross_section,
                wire_type=wire_type,
                insulation_material=insulation_material,
                fire_resistant=fire_resistant,
                length=length,
            ))

        else:  # Кабель
            cable_mark = name_or_mark
            cable_stock.append(CableStock(
                id=item_id,
                cable_mark=cable_mark,
                length=length,
            ))

    return raw_wires, insulated, cable_stock


def _parse_drums(ws) -> Tuple[List[CoreDrumCapacity], List[CableDrumCapacity]]:
    """
    Лист '4. Барабаны'.

    Секция А (жилы):
      Шапка: 'Сечение, мм²' | 'Индекс' | 'Б-400, м' | 'Б-630, м' | ...
      Данные: '2,5' | 'ок' | 2500 | 4500 | ...
      Ключ CoreDrumCapacity.wire_key = cross_section + wire_type, например '2,5ок'.

    Секция Б (кабели):
      Шапка: 'Марка кабеля' | '№10, м' | '№12, м' | ...
      Данные: '<марка>' | 1000 | 2000 | ...
    """
    core_caps: List[CoreDrumCapacity] = []
    cable_caps: List[CableDrumCapacity] = []

    rows = list(ws.iter_rows(values_only=True))

    def _find_header_row(keyword: str) -> Optional[int]:
        for i, row in enumerate(rows):
            if row and _str(row[0]) == keyword:
                return i
        return None

    def _read_core_section(header_idx: int) -> Tuple[List[str], List[tuple]]:
        """
        Читает секцию А: 2 ключевых столбца (Сечение + Индекс), затем барабаны.
        Возвращает: (drum_names, [(wire_key, [cap1, cap2, ...])]).
        """
        header_row = rows[header_idx]
        drum_names = []
        for col_val in header_row[2:]:  # барабаны начинаются с 3-й колонки
            s = _str(col_val)
            if not s or '←' in s or 'Добавляйте' in s:
                break
            drum_names.append(_drum_type_name(s))

        data_rows = []
        for row in rows[header_idx + 1:]:
            if not row:
                break
            cross_raw = _str(row[0])
            # Остановиться на подсказке ('←') или пустой строке
            if not cross_raw or cross_raw.startswith('←') or not cross_raw[0].isdigit():
                break
            cross = _cross_section_str(row[0])
            wtype = _str(row[1]) if len(row) > 1 else ''
            wire_key = f'{cross}{wtype}'
            capacities = [_float(row[j + 2]) for j in range(len(drum_names))]
            data_rows.append((wire_key, capacities))

        return drum_names, data_rows

    def _read_cable_section(header_idx: int) -> Tuple[List[str], List[tuple]]:
        """
        Читает секцию Б: 1 ключевой столбец (Марка кабеля), затем барабаны.
        Возвращает: (drum_names, [(cable_mark, [cap1, cap2, ...])]).
        """
        header_row = rows[header_idx]
        drum_names = []
        for col_val in header_row[1:]:
            s = _str(col_val)
            if not s or '←' in s or 'Добавляйте' in s:
                break
            drum_names.append(_drum_type_name(s))

        data_rows = []
        for row in rows[header_idx + 1:]:
            if not row:
                break
            key = _str(row[0])
            if not key:
                break
            capacities = [_float(row[j + 1]) for j in range(len(drum_names))]
            data_rows.append((key, capacities))

        return drum_names, data_rows

    # ── Секция А: жилы ──────────────────────────────────────────────
    idx_a = _find_header_row('Сечение, мм²')
    if idx_a is not None:
        drum_names, data_rows = _read_core_section(idx_a)
        for wire_key, capacities in data_rows:
            drum_types = [
                DrumType(name=dn, capacity=cap)
                for dn, cap in zip(drum_names, capacities)
                if cap > 0
            ]
            drum_types.sort(key=lambda d: d.capacity)
            core_caps.append(CoreDrumCapacity(
                wire_key=wire_key,
                drum_types=drum_types,
            ))

    # ── Секция Б: кабели ────────────────────────────────────────────
    idx_b = _find_header_row('Марка кабеля')
    if idx_b is not None:
        drum_names, data_rows = _read_cable_section(idx_b)
        for cable_mark, capacities in data_rows:
            drum_types = [
                DrumType(name=dn, capacity=cap)
                for dn, cap in zip(drum_names, capacities)
                if cap > 0
            ]
            drum_types.sort(key=lambda d: d.capacity)
            cable_caps.append(CableDrumCapacity(
                cable_mark=cable_mark,
                drum_types=drum_types,
            ))

    return core_caps, cable_caps


def _parse_params(ws) -> ProcessParams:
    """
    Лист '5. Параметры'.
    Колонки: A=Параметр, B=Значение
    """
    params = ProcessParams()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        key = _str(row[0]).lower()
        val = row[1] if len(row) > 1 else None

        if 'изолирован' in key:
            params.max_insulation_run = _float(val, params.max_insulation_run)
        elif 'скрутк' in key or 'партии' in key:
            params.max_twisting_run = _float(val, params.max_twisting_run)
        elif 'строительн' in key or 'минимальн' in key:
            params.min_construction_length = _float(val, params.min_construction_length)
        elif 'спайк' in key:
            s = _str(val).lower()
            params.allow_splicing = s in ('да', 'yes', '1', 'true')
        elif 'несколько' in key or 'мультисегм' in key or 'multi' in key:
            s = _str(val).lower()
            params.allow_multi_segment_drum = s in ('да', 'yes', '1', 'true')
        elif 'стратег' in key:
            params.strategy = _str(val)
        elif 'заправк' in key or 'startup' in key:
            params.insulation_startup_loss_m = _float(val, params.insulation_startup_loss_m)
        elif ('допуск' in key or 'запас' in key) and ('торц' in key or 'обрезк' in key or 'длин' in key):
            params.length_tolerance_m = _float(val, params.length_tolerance_m)
        elif 'порядок' in key and ('журнал' in key or 'кабельн' in key):
            s = _str(val).lower()
            params.keep_journal_order = s in ('да', 'yes', '1', 'true')
        elif 'отход' in key or 'порог' in key:
            params.waste_warning_threshold_m = _float(val, params.waste_warning_threshold_m)

    return params


# ─── Главная функция парсинга ────────────────────────────────────────

def parse_input(filepath: str):
    """
    Читает input.xlsx и возвращает все объекты для планирования.

    Возвращает:
        orders, raw_wires, insulated_cores, cable_stock,
        core_drum_caps, cable_drum_caps, params
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    def _sheet(keyword: str):
        for name in wb.sheetnames:
            if keyword.lower() in name.lower():
                return wb[name]
        raise KeyError(f'Лист с "{keyword}" не найден в файле {filepath}')

    orders = _parse_orders(_sheet('Заказы'))
    composition = _parse_composition(_sheet('Состав'))
    raw_wires, insulated_cores, cable_stock = _parse_pf(_sheet('П-Ф'))
    core_drum_caps, cable_drum_caps = _parse_drums(_sheet('Барабан'))
    params = _parse_params(_sheet('Параметр'))

    # Обогащаем заказы атрибутами жил из листа состава
    warnings = []
    for order in orders:
        if order.mark in composition:
            cross_section, wire_type, fire_resistant, insulation_material, colors = composition[order.mark]
            order.cross_section = cross_section
            order.wire_type = wire_type
            order.fire_resistant = fire_resistant
            order.insulation_material = insulation_material
            order.colors = colors
        else:
            warnings.append(
                f'⚠ Марка "{order.mark}" не найдена в листе "Состав кабелей" — '
                f'цвета и сечение не определены.'
            )

    if warnings:
        print('\n'.join(warnings))

    return orders, raw_wires, insulated_cores, cable_stock, core_drum_caps, cable_drum_caps, params
