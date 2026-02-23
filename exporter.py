"""
Экспорт результатов планирования в Excel.
"""
from __future__ import annotations
from collections import defaultdict
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models import PlanResult, InsulationRun, InsulatedCoreUse


# ─── Стили (те же, что в шаблоне входного файла) ─────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def _thin_border() -> Border:
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font()  -> Font: return Font(name='Calibri', bold=True, color='FFFFFF', size=10)
def _body_font() -> Font: return Font(name='Calibri', size=10)
def _note_font() -> Font: return Font(name='Calibri', italic=True, color='595959', size=9)
def _err_font()  -> Font: return Font(name='Calibri', bold=True, color='C00000', size=10)

C_HDR       = '2E75B6'
C_EVEN      = 'DEEAF1'
C_ODD       = 'FFFFFF'
C_WARN_ROW  = 'FFF2CC'
C_ERR_ROW   = 'FCE4D6'
C_SEC_TITLE = 'BDD7EE'
C_STOCK_ROW = 'E2EFDA'    # строки из склада


# ─── Вспомогательные функции листа ───────────────────────────────────

def _autowidth(ws, max_col: int = 20, cap: int = 60):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            # MergedCell не имеет column_letter — берём из первой обычной ячейки
            if col_letter is None and hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
            if not hasattr(cell, 'column_letter'):
                continue
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 2, cap)


def _write_header(ws, row: int, headers: list, col_start: int = 1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = _hdr_font()
        c.fill = _fill(C_HDR)
        c.border = _thin_border()
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30


def _write_row(ws, row: int, values: list, col_start: int = 1,
               even: bool = False, bg: str = None):
    bg_color = bg if bg else (C_EVEN if even else C_ODD)
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=val)
        c.font = _body_font()
        c.fill = _fill(bg_color)
        c.border = _thin_border()
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _section_title(ws, row: int, text: str, n_cols: int):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Calibri', bold=True, color='1F4E79', size=11)
    c.fill = _fill(C_SEC_TITLE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    ws.row_dimensions[row].height = 18


def _color_fill_for(color: str) -> str:
    """Цвет заливки строки по названию жилы."""
    COLOR_MAP = {
        'синяя':          'ADD8E6',
        'желто-зеленая':  'FFFF99',
        'желто-зелёная':  'FFFF99',
        'ж/з':            'FFFF99',
        'натуральная':    'F5F5DC',
        'чёрная':         'D3D3D3',
        'черная':         'D3D3D3',
        'коричневая':     'DEB887',
        'серая':          'E8E8E8',
    }
    return COLOR_MAP.get(color.lower(), C_ODD)


# ═══════════════════════════════════════════════════════════════════════
# Вспомогательные данные: реестр П/Ф с порядковыми номерами
# ═══════════════════════════════════════════════════════════════════════

def _build_pf_data(result: PlanResult):
    """
    Строит нумерованный реестр полуфабрикатов.

    Поддерживает несколько катушек на один цвет в рамках одной партии
    (multi-spool режим): каждая физическая катушка получает свой номер П/Ф.

    Порядок: по партиям → по позиции цвета → по spool_index.

    Возвращает:
        pf_map       — dict[(batch_id, color, spool_index)] → номер (int)
        pf_registry  — list of dict, одна запись = один П/Ф
    """
    ins_run_map: dict = {}
    for run in result.insulation_runs:
        ins_run_map[(run.for_batch_id, run.color)] = run

    # Для склада: (batch_id, color) → sorted list[InsulatedCoreUse]
    ins_use_map_multi: dict = {}
    for use in result.insulated_core_uses:
        key = (use.for_batch_id, use.color)
        ins_use_map_multi.setdefault(key, []).append(use)
    for key in ins_use_map_multi:
        ins_use_map_multi[key].sort(key=lambda u: u.spool_index)

    pf_map: dict = {}       # (batch_id, color, spool_index) → pf_num
    pf_registry: list = []

    n = 1
    for batch in result.batches:
        for color in batch.colors:
            run  = ins_run_map.get((batch.id, color))
            uses = ins_use_map_multi.get((batch.id, color), [])

            if run:
                # Один прогон изолирования = одна катушка
                pf_map[(batch.id, color, 1)] = n
                mat = run.insulation_material + (f'+{run.fire_resistant}' if run.fire_resistant else '')
                pf_registry.append({
                    'num':         n,
                    'status':      'Изолировать',
                    'batch_id':    batch.id,
                    'cable_mark':  batch.cable_mark,
                    'color':       color,
                    'wire_key':    run.wire_key,
                    'material':    mat,
                    'length':      int(run.length),
                    'source':      run.source_name,
                    'drum_type':   run.drum_type,
                    'run_id':      run.id,
                    'note':        f'Написать «П/Ф №{n}» на мотке',
                })
                n += 1

            elif uses:
                for use in uses:
                    pf_map[(batch.id, color, use.spool_index)] = n
                    segs_str = '+'.join(str(int(batch.segments[j]))
                                       for j in use.covered_segments)
                    pf_registry.append({
                        'num':         n,
                        'status':      'Взять со склада',
                        'batch_id':    batch.id,
                        'cable_mark':  batch.cable_mark,
                        'color':       color,
                        'wire_key':    use.wire_key,
                        'material':    '(склад)',
                        'length':      int(use.length),
                        'source':      use.source_name,
                        'drum_type':   '(готовая)',
                        'run_id':      use.id,
                        'note':        (f'Взять {int(use.length)} м, остаток: {int(use.remainder)} м'
                                        + (f' → сегменты [{segs_str}]' if len(uses) > 1 else '')),
                    })
                    n += 1

            else:
                pf_map[(batch.id, color, 1)] = n
                pf_registry.append({
                    'num':         n,
                    'status':      '? Не определено',
                    'batch_id':    batch.id,
                    'cable_mark':  batch.cable_mark,
                    'color':       color,
                    'wire_key':    '?',
                    'material':    '?',
                    'length':      0,
                    'source':      '?',
                    'drum_type':   '?',
                    'run_id':      '?',
                    'note':        '',
                })
                n += 1

    return pf_map, pf_registry


def _rebuild_tpzh_balance(result: PlanResult) -> dict:
    """
    Восстанавливает исходный баланс каждого барабана ТПЖ из списка
    insulation_runs и remaining_raw_wires.
    Возвращает dict[source_id] → original_length.
    """
    used_per_drum: dict = defaultdict(float)
    for run in result.insulation_runs:
        # П1+П9: используем raw_wire_consumed (= length + потери на заправку),
        # чтобы исходный баланс ТПЖ восстанавливался правильно.
        consumed = run.raw_wire_consumed if run.raw_wire_consumed > 0 else run.length
        used_per_drum[run.source_id] += consumed

    remain_per_drum: dict = {}
    for rw in result.remaining_raw_wires:
        remain_per_drum[rw.id] = rw.available

    all_ids = set(used_per_drum) | set(remain_per_drum)
    original: dict = {}
    for sid in all_ids:
        original[sid] = used_per_drum.get(sid, 0.0) + remain_per_drum.get(sid, 0.0)
    return original


# ═══════════════════════════════════════════════════════════════════════
# Лист 1: Сводка
# ═══════════════════════════════════════════════════════════════════════

def _sheet_summary(wb: openpyxl.Workbook, result: PlanResult):
    ws = wb.active
    ws.title = '1. Сводка'

    row = 1
    ws.cell(row=row, column=1, value='СВОДКА ПЛАНИРОВАНИЯ').font = Font(
        name='Calibri', bold=True, color='1F4E79', size=14)
    ws.merge_cells(f'A{row}:F{row}')
    row += 2

    summary_rows = [
        ('Заказов',                len(result.orders)),
        ('Партий скрутки',         len(result.batches)),
        ('Прогонов изолирования',  len(result.insulation_runs)),
        ('Исп. жил со склада',     len(result.insulated_core_uses)),
        ('Исп. кабеля со склада',  len(result.cable_stock_uses)),
        ('Выходных барабанов',     len(result.drum_assignments)),
        ('Ошибок',                 len(result.errors)),
        ('Предупреждений',         len(result.warnings)),
    ]
    for label, val in summary_rows:
        ws.cell(row=row, column=1, value=label).font = _body_font()
        c = ws.cell(row=row, column=2, value=val)
        c.font = Font(name='Calibri', bold=True, size=10)
        row += 1

    _autowidth(ws)


# ═══════════════════════════════════════════════════════════════════════
# Лист 2: Ведомость полуфабрикатов (реестр жил с номерами)
# ═══════════════════════════════════════════════════════════════════════

def _sheet_pf_registry(wb: openpyxl.Workbook, result: PlanResult, pf_registry: list):
    """
    Ведомость всех изолированных жил, участвующих в производственной программе.
    Каждой жиле присвоен сквозной порядковый номер, который:
      • изоляторщик пишет на мотке после прогона;
      • скрутчик использует для комплектации партии.
    """
    ws = wb.create_sheet('2. Ведомость ПФ')

    headers = [
        '№ П/Ф', 'Статус', 'Партия скрутки', 'Марка кабеля',
        'Цвет жилы', 'Тип жилы', 'Материал/FR',
        'Длина, м', 'Источник (ТПЖ-барабан / бухта)', 'Принять на (моток)',
        'Примечание',
    ]
    _section_title(ws, 1, 'ВЕДОМОСТЬ ПОЛУФАБРИКАТОВ — нумерованные жилы для скрутки', len(headers))
    _write_header(ws, 2, headers)

    C_INS  = 'FFF2CC'   # жёлтый — требует изолирования
    C_STK  = 'E2EFDA'   # зелёный — со склада
    C_ERR  = 'FCE4D6'   # красный — не определено

    status_color = {
        'Изолировать':    C_INS,
        'Взять со склада': C_STK,
        '❌ Не определено': C_ERR,
    }

    row = 3
    for entry in pf_registry:
        bg = _color_fill_for(entry['color'])
        # Строки «со склада» — немного зеленоватее
        if entry['status'] == 'Взять со склада':
            bg = C_STK
        elif entry['status'].startswith('❌'):
            bg = C_ERR

        _write_row(ws, row, [
            entry['num'],
            entry['status'],
            entry['batch_id'],
            entry['cable_mark'],
            entry['color'],
            entry['wire_key'],
            entry['material'],
            entry['length'],
            entry['source'],
            entry['drum_type'],
            entry['note'],
        ], bg=bg)

        # № П/Ф — жирный, крупнее
        c_num = ws.cell(row=row, column=1)
        c_num.font = Font(name='Calibri', bold=True, size=11)
        c_num.alignment = Alignment(horizontal='center', vertical='center')

        row += 1

    ws.freeze_panes = 'A3'
    ws.column_dimensions['A'].width = 8
    _autowidth(ws)


# ═══════════════════════════════════════════════════════════════════════
# Лист 3: Инструкция изолирования
# ═══════════════════════════════════════════════════════════════════════

def _sheet_insulation(wb: openpyxl.Workbook, result: PlanResult, pf_map: dict):
    ws = wb.create_sheet('3. Изолирование')

    headers = [
        '№ П/Ф', 'Партия скрутки', 'Марка кабеля',
        'Цвет жилы', 'Тип жилы', 'Материал/FR',
        'Источник ТПЖ', 'Было на ТПЖ, м', 'Длина прогона, м', 'Снять с барабана, м',
        'Остаток ТПЖ, м', 'Принять на (моток)',
    ]
    _section_title(ws, 1, '📐  ИНСТРУКЦИЯ ИЗОЛИРОВАНИЯ', len(headers))
    _write_header(ws, 2, headers)

    batch_mark: dict = {b.id: b.cable_mark for b in result.batches}

    # Восстанавливаем исходный баланс каждого ТПЖ-барабана
    tpzh_original = _rebuild_tpzh_balance(result)
    tpzh_current: dict = dict(tpzh_original)   # рабочий бегущий баланс

    # Сортируем прогоны в том же порядке, что и реестр: по партии, затем по позиции цвета
    batch_order = {b.id: i for i, b in enumerate(result.batches)}
    batch_color_pos = {}
    for b in result.batches:
        for pos, color in enumerate(b.colors):
            batch_color_pos[(b.id, color)] = pos

    sorted_runs = sorted(
        result.insulation_runs,
        key=lambda r: (batch_order.get(r.for_batch_id, 999),
                       batch_color_pos.get((r.for_batch_id, r.color), 999))
    )
    sorted_uses = sorted(
        result.insulated_core_uses,
        key=lambda u: (batch_order.get(u.for_batch_id, 999),
                       batch_color_pos.get((u.for_batch_id, u.color), 999),
                       u.spool_index)
    )

    row = 3
    for run in sorted_runs:
        pf_num = pf_map.get((run.for_batch_id, run.color, 1), '?')
        mark = batch_mark.get(run.for_batch_id, '')
        consumed = run.raw_wire_consumed if run.raw_wire_consumed > 0 else run.length
        before = round(tpzh_current.get(run.source_id, 0))
        tpzh_current[run.source_id] = tpzh_current.get(run.source_id, 0) - consumed
        after = round(tpzh_current[run.source_id])
        mat = run.insulation_material + (f'+{run.fire_resistant}' if run.fire_resistant else '')
        bg = _color_fill_for(run.color)

        _write_row(ws, row, [
            pf_num,
            run.for_batch_id,
            mark,
            run.color,
            run.wire_key,
            mat,
            run.source_name,
            before,
            int(run.length),
            int(consumed),
            after,
            run.drum_type,
        ], bg=bg)
        # № — жирный
        ws.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=11)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    # Строки «из склада изолированных»
    if sorted_uses:
        _section_title(ws, row, '  Из склада изолированных жил (готовые — взять без изолирования)', len(headers))
        row += 1
        for use in sorted_uses:
            pf_num = pf_map.get((use.for_batch_id, use.color, use.spool_index), '?')
            mark = batch_mark.get(use.for_batch_id, '')
            bg = C_STOCK_ROW
            _write_row(ws, row, [
                pf_num,
                use.for_batch_id,
                mark,
                use.color,
                use.wire_key,
                '(склад)',
                use.source_name,
                int(use.length + use.remainder),
                int(use.length),
                int(use.remainder),
                '(уже готовая)',
            ], bg=bg)
            ws.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=11)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            row += 1

    ws.freeze_panes = 'A3'
    _autowidth(ws)


# ═══════════════════════════════════════════════════════════════════════
# Лист 4: Таблица заправок скрутки
# ═══════════════════════════════════════════════════════════════════════

def _sheet_twisting(wb: openpyxl.Workbook, result: PlanResult, pf_map: dict):
    """
    ТАБЛИЦА ЗАПРАВОК — для каждой партии скрутки:
      • строка заправки: П/Ф-номер + исходная длина по каждой позиции
      • строки шагов: для каждого отрезка журнала — длина до и после по каждой позиции
    """
    ws = wb.create_sheet('4. Скрутка')

    # Карты для быстрого поиска источника жилы
    ins_run_map: dict = {}
    for run in result.insulation_runs:
        ins_run_map[(run.for_batch_id, run.color)] = run

    # (batch_id, color) → первая катушка (spool_index минимальный)
    ins_use_first: dict = {}
    for use in result.insulated_core_uses:
        key = (use.for_batch_id, use.color)
        if key not in ins_use_first or use.spool_index < ins_use_first[key].spool_index:
            ins_use_first[key] = use

    max_colors = max((len(b.colors) for b in result.batches), default=0)
    N_FIXED = 3   # фиксированные колонки: Шаг | Длина, м | Нарастающим, м
    total_cols = N_FIXED + max_colors

    _section_title(ws, 1, 'ТАБЛИЦА ЗАПРАВОК — инструкция скрутки', total_cols)

    row = 2
    for batch in result.batches:
        n_colors = len(batch.colors)
        fr_info  = f'+{batch.fire_resistant}' if batch.fire_resistant else ''
        mat_info = batch.insulation_material + fr_info

        # ── Исходная длина первой катушки на каждую позицию ─────────────
        # (для строки ЗАПРАВКА; multi-spool — показывает первую катушку)
        spool_len: dict = {}
        for color in batch.colors:
            key = (batch.id, color)
            run  = ins_run_map.get(key)
            use0 = ins_use_first.get(key)
            if run:
                spool_len[color] = run.length
            elif use0:
                spool_len[color] = use0.length
            else:
                spool_len[color] = 0.0

        # ── Заголовок партии ─────────────────────────────────────────────
        batch_title = (
            f'{batch.id}   |   {batch.cable_mark}   |   {batch.wire_key}'
            + (f'   |   {mat_info}' if mat_info else '')
            + f'   |   Итого: {int(batch.total_length)} м'
            + f'   |   Отрезков: {len(batch.segments)}'
        )
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=total_cols)
        c = ws.cell(row=row, column=1, value=batch_title)
        c.font      = Font(name='Calibri', bold=True, color='1F4E79', size=11)
        c.fill      = _fill(C_SEC_TITLE)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 18
        row += 1

        # ── Заголовки колонок ────────────────────────────────────────────
        color_hdrs = [
            f'Поз.{i+1}  {batch.colors[i]}' for i in range(n_colors)
        ] + [''] * (max_colors - n_colors)
        step_hdrs = ['Шаг', 'Длина, м', 'Нарастающим, м'] + color_hdrs
        _write_header(ws, row, step_hdrs)
        row += 1

        # ── Строка заправки ──────────────────────────────────────────────
        setup_vals = ['ЗАПРАВКА', '—', '—']
        for color in batch.colors:
            # Показываем первую катушку (spool_index=1)
            pf_num  = pf_map.get((batch.id, color, 1), '?')
            initial = int(round(spool_len.get(color, 0)))
            setup_vals.append(f'П/Ф №{pf_num}\n{initial} м')
        setup_vals += [''] * (max_colors - n_colors)

        for col_i, val in enumerate(setup_vals):
            c = ws.cell(row=row, column=col_i + 1, value=val)
            color_i = col_i - N_FIXED
            if 0 <= color_i < n_colors:
                bg = _color_fill_for(batch.colors[color_i])
            else:
                bg = C_ODD
            c.fill      = _fill(bg)
            c.font      = Font(name='Calibri', bold=True, size=10)
            c.border    = _thin_border()
            c.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True)
        ws.row_dimensions[row].height = 40
        row += 1

        # ── Таблица seg_idx → InsulatedCoreUse для этой партии ──────────
        # Чтобы знать, какая катушка стоит на каждом шаге.
        # Для InsulationRun: одна катушка на все сегменты (spool_index=1).
        seg_to_use: dict = {}    # (color, seg_idx) → InsulatedCoreUse or None
        use_balance: dict = {}   # use.id → текущий остаток (начало = use.length)

        for use in result.insulated_core_uses:
            if use.for_batch_id != batch.id:
                continue
            seg_indices = use.covered_segments if use.covered_segments else list(range(len(batch.segments)))
            for si in seg_indices:
                seg_to_use[(use.color, si)] = use
            if use.id not in use_balance:
                use_balance[use.id] = use.length   # начальный остаток этой катушки

        # Для InsulationRun — смоделируем единственный «use» через spool_len
        # (баланс уже хранится в spool_len; используем отдельный словарь run_balance)
        run_balance: dict = {}
        for color in batch.colors:
            run = ins_run_map.get((batch.id, color))
            if run:
                run_balance[color] = spool_len.get(color, 0.0)

        # ── Строки шагов (по отрезкам журнала) ──────────────────────────
        cumulative = 0.0

        for step_i, seg in enumerate(batch.segments):
            cumulative += seg
            row_vals = [step_i + 1, int(round(seg)), int(round(cumulative))]

            for color in batch.colors:
                run = ins_run_map.get((batch.id, color))
                use = seg_to_use.get((color, step_i))

                if run:
                    # Один прогон — одна катушка на всю партию
                    pf_num = pf_map.get((batch.id, color, 1), '?')
                    before = int(round(run_balance[color]))
                    run_balance[color] -= seg
                    after  = int(round(run_balance[color]))
                    row_vals.append(f'П/Ф №{pf_num}: {before}м\n(ост: {after}м)')

                elif use:
                    pf_num = pf_map.get((batch.id, color, use.spool_index), '?')
                    before = int(round(use_balance[use.id]))
                    use_balance[use.id] -= seg
                    after  = int(round(use_balance[use.id]))

                    # Отмечаем смену катушки на следующем шаге
                    next_use = seg_to_use.get((color, step_i + 1)) if step_i + 1 < len(batch.segments) else None
                    spool_change = (next_use is not None and next_use.id != use.id)
                    suffix = '\n↓ СМЕНА КАТУШКИ' if spool_change else ''

                    row_vals.append(f'П/Ф №{pf_num}: {before}м\n(ост: {after}м){suffix}')

                else:
                    row_vals.append('—')
            row_vals += [''] * (max_colors - n_colors)

            for col_i, val in enumerate(row_vals):
                c = ws.cell(row=row, column=col_i + 1, value=val)
                color_i = col_i - N_FIXED
                if 0 <= color_i < n_colors:
                    bg = _color_fill_for(batch.colors[color_i])
                else:
                    bg = C_EVEN if step_i % 2 else C_ODD
                c.fill      = _fill(bg)
                c.font      = _body_font()
                c.border    = _thin_border()
                c.alignment = Alignment(horizontal='center', vertical='center',
                                        wrap_text=True)
            ws.row_dimensions[row].height = 36
            row += 1

        # Пустая строка-разделитель между партиями
        row += 1

    # ── Ширины колонок ───────────────────────────────────────────────────
    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 16
    for i in range(max_colors):
        col_letter = get_column_letter(N_FIXED + 1 + i)
        ws.column_dimensions[col_letter].width = 24


# ═══════════════════════════════════════════════════════════════════════
# Лист 4: Инструкция намотки / выходные барабаны
# ═══════════════════════════════════════════════════════════════════════

def _sheet_drums(wb: openpyxl.Workbook, result: PlanResult):
    ws = wb.create_sheet('5. Барабаны (выход)')

    headers = [
        '№ барабана', 'Марка кабеля', 'Тип барабана',
        'Ёмкость, м', 'Отрезки журнала (м)',
        'Суммарно, м', 'Загрузка, %', 'Источник',
    ]
    _section_title(ws, 1, '🛢  ВЫХОДНЫЕ БАРАБАНЫ — ИНСТРУКЦИЯ НАМОТКИ', len(headers))
    _write_header(ws, 2, headers)

    row = 3
    for i, da in enumerate(result.drum_assignments):
        segs_str = ', '.join(str(int(s)) for s in da.segments)
        pct = round(da.total_length / da.drum_capacity * 100, 1) if da.drum_capacity else 0
        bg = C_STOCK_ROW if da.source == 'склад' else (C_EVEN if i % 2 else C_ODD)
        _write_row(ws, row, [
            i + 1,
            da.cable_mark,
            da.drum_type,
            int(da.drum_capacity),
            segs_str,
            int(da.total_length),
            f'{pct}%',
            da.source,
        ], bg=bg)
        row += 1

    ws.freeze_panes = 'A3'
    _autowidth(ws)


# ═══════════════════════════════════════════════════════════════════════
# Лист 5: Остатки ТПЖ
# ═══════════════════════════════════════════════════════════════════════

def _sheet_raw_remains(wb: openpyxl.Workbook, result: PlanResult):
    ws = wb.create_sheet('6. Остатки ТПЖ')

    headers = ['Барабан (ID)', 'Наименование', 'Тип жилы', 'Всего, м', 'Использовано, м', 'Остаток, м', 'Использование, %']
    _section_title(ws, 1, '📦  ОСТАТКИ НЕИЗОЛИРОВАННОЙ ЖИЛЫ (ТПЖ)', len(headers))
    _write_header(ws, 2, headers)

    row = 3
    for i, rw in enumerate(result.remaining_raw_wires):
        used = rw.length - rw.available
        pct = round(used / rw.length * 100, 1) if rw.length else 0
        _write_row(ws, row, [
            rw.id, rw.name, rw.wire_key,
            int(rw.length), int(used), int(rw.available), f'{pct}%',
        ], even=(i % 2 == 1))
        row += 1

    _autowidth(ws)


# ═══════════════════════════════════════════════════════════════════════
# Лист 6: Остатки изолированных жил и склад кабеля
# ═══════════════════════════════════════════════════════════════════════

def _sheet_ins_remains(wb: openpyxl.Workbook, result: PlanResult):
    ws = wb.create_sheet('7. Остатки жил и кабеля')

    headers = ['ID', 'Тип', 'Наименование / Марка', 'Цвет / Сечение', 'Остаток, м']
    _section_title(ws, 1, '📦  ОСТАТКИ ГОТОВЫХ ЖИЛ И КАБЕЛЯ НА СКЛАДЕ', len(headers))
    _write_header(ws, 2, headers)

    row = 3
    for i, ins in enumerate(result.remaining_insulated):
        _write_row(ws, row, [
            ins.id, 'Изолированная', ins.name, ins.color, int(ins.available),
        ], even=(i % 2 == 1), bg=_color_fill_for(ins.color))
        row += 1

    if result.remaining_cable_stock:
        _section_title(ws, row, '  Готовый кабель', len(headers))
        row += 1
        for i, cab in enumerate(result.remaining_cable_stock):
            _write_row(ws, row, [
                cab.id, 'Кабель', cab.cable_mark, '—', int(cab.available),
            ], even=(i % 2 == 1), bg=C_STOCK_ROW)
            row += 1

    _autowidth(ws)


# ═══════════════════════════════════════════════════════════════════════
# Лист 7: Ошибки и предупреждения
# ═══════════════════════════════════════════════════════════════════════

def _sheet_errors(wb: openpyxl.Workbook, result: PlanResult):
    if not result.errors and not result.warnings:
        return

    ws = wb.create_sheet('8. Ошибки')
    ws.cell(row=1, column=1, value='ОШИБКИ И ПРЕДУПРЕЖДЕНИЯ').font = Font(
        name='Calibri', bold=True, color='C00000', size=13)
    ws.merge_cells('A1:B1')

    row = 2
    for err in result.errors:
        c = ws.cell(row=row, column=1, value=err)
        c.font = _err_font()
        c.fill = _fill(C_ERR_ROW)
        c.border = _thin_border()
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

    for warn in result.warnings:
        c = ws.cell(row=row, column=1, value=warn)
        c.font = _body_font()
        c.fill = _fill(C_WARN_ROW)
        c.border = _thin_border()
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

    ws.column_dimensions['A'].width = 100


# ═══════════════════════════════════════════════════════════════════════
# Главная функция
# ═══════════════════════════════════════════════════════════════════════

def export(filepath: str, result: PlanResult):
    """Экспортирует результат планирования в Excel-файл."""
    wb = openpyxl.Workbook()

    # Строим реестр П/Ф один раз — используется в нескольких листах
    pf_map, pf_registry = _build_pf_data(result)

    _sheet_summary(wb, result)
    _sheet_pf_registry(wb, result, pf_registry)
    _sheet_insulation(wb, result, pf_map)
    _sheet_twisting(wb, result, pf_map)
    _sheet_drums(wb, result)
    _sheet_raw_remains(wb, result)
    _sheet_ins_remains(wb, result)
    _sheet_errors(wb, result)

    wb.save(filepath)
    print(f'Результат сохранён: {filepath}')
