import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────
def fill(hex_): return PatternFill(start_color=hex_, end_color=hex_, fill_type='solid')
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'),  bottom=Side(style='thin'))
no_border = Border()

BLUE_HDR   = 'BDD7EE'
GRAY_HDR   = 'D6D6D6'
YELLOW     = 'FFFACD'
LIGHT_BLUE = 'DDEEFF'
GRAY_CELL  = 'EBEBEB'
BROWN      = 'F5E6D3'
CREAM      = 'F8F8F0'
PAIR_CLR   = 'E8F5E9'
PAIR_ID    = 'D6F5D6'
WARN_CLR   = 'FFF3CD'
PARAM_HDR  = 'E2EFDA'

def col(ws, i, w):
    ws.column_dimensions[get_column_letter(i)].width = w

def set_hdr(ws, row, c, val, bg=BLUE_HDR, bold=True, wrap=True, align='center'):
    cell = ws.cell(row=row, column=c, value=val)
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = thin
    return cell

def set_data(ws, row, c, val, bg=None, bold=False, align='left'):
    cell = ws.cell(row=row, column=c, value=val)
    if bg: cell.fill = fill(bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = thin
    return cell

def note_row(ws, row, text, ncols, bg=WARN_CLR):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill(bg)
    cell.font = Font(italic=True, size=9, color='7B5A00')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = no_border
    ws.row_dimensions[row].height = 36

def title_row(ws, row, text, ncols, bg='1F4E79'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill(bg)
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 24


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: ЗАКАЗЫ
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '1. Заказы'

title_row(ws1, 1, '📋  ЗАКАЗЫ НА КАБЕЛЬ', 5)
note_row(ws1, 2,
    '⚠  Марка кабеля — совпадает с листом «2. Состав кабелей».  '
    'Гибкий = ДА → алгоритм может увеличить длину жилы для поглощения остатков барабана.  '
    'Кабельный журнал — строительные длины через запятую (напр.: 2000, 1500, 300); '
    'если пусто — берётся поле «Длина, м» целиком как одна строительная длина.',
    5)

h1 = ['Марка кабеля', 'Длина, м', 'Гибкий\n(ДА/НЕТ)', 'Кабельный журнал\n(строит. длины через запятую)', 'Примечание']
w1 = [52, 12, 12, 42, 30]
for i, (h, w) in enumerate(zip(h1, w1), 1):
    set_hdr(ws1, 3, i, h)
    col(ws1, i, w)

orders = [
    # --- Силовые 5ж ---
    ('п.10 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1',  506,  'НЕТ', '',    '460м × 1.10'),
    ('п.30 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1',  473,  'НЕТ', '',    '430м × 1.10'),
    ('п.31 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1',  737,  'НЕТ', '',    '670м × 1.10'),
    ('п.37 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1',  869,  'НЕТ', '',    '790м × 1.10'),
    ('п.42 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1',  385,  'НЕТ', '',    '350м × 1.10'),
    ('п.51 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1',  776,  'НЕТ', '',    '705м × 1.10'),
    ('п.53 Вз-ВВГнг(А)-LS 5х16мк(N,PE)-1',   440,  'НЕТ', '',    '400м × 1.10'),
    ('п.56 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1',  1139, 'НЕТ', '',    '1035м × 1.10'),
    ('п.88 Вз-ВВГнг(А)-LS 5х16мк(N,PE)-1',   132,  'НЕТ', '',    '120м × 1.10'),
    # --- Силовые 4ж ---
    ('п.14/1 Вз-ВБШвнг(PE)-LS 4х16мк(N)-1',  1000, 'ДА',  '',    'гибкий — поглощает остатки'),
    ('п.14/2 Вз-ВБШвнг(PE)-LS 4х16мк(N)-1',  1000, 'ДА',  '',    ''),
    ('п.14/3 Вз-ВБШвнг(PE)-LS 4х16мк(N)-1',  1000, 'ДА',  '',    ''),
    ('п.14/4 Вз-ВБШвнг(PE)-LS 4х16мк(N)-1',  1000, 'ДА',  '',    ''),
    # --- Парные ---
    ('п.20 ТПП 4х2х1',   1200, 'НЕТ', '600, 600', 'парный кабель — 4 пары × 2 жилы'),
    ('п.21 ТПП 8х2х0.5',  800, 'НЕТ', '400, 400', 'парный — 8 пар × 2 жилы'),
]

row = 4
for rec in orders:
    is_pair = 'ТПП' in rec[0] or 'пар' in rec[4].lower()
    bg_row = PAIR_CLR if is_pair else None
    for ci, v in enumerate(rec, 1):
        bg = bg_row
        if ci == 3:
            bg = 'E8F8E8' if v == 'ДА' else ('E8F4FD' if not is_pair else PAIR_CLR)
        set_data(ws1, row, ci, v if v != '' else None,
                 bg=bg, align='center' if ci in (2, 3) else 'left')
    row += 1

ws1.freeze_panes = 'A4'


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: СОСТАВ КАБЕЛЕЙ
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('2. Состав кабелей')

title_row(ws2, 1, '🎨  СОСТАВ КАБЕЛЕЙ — жилы, цвета и номера пар', 14)
note_row(ws2, 2,
    '⚠  Тип «Силовой»: Ид. 1..N — цвета жил (ж/з, синий, чёрный, коричневый, натуральный).  '
    'Тип «Парный»: Ид. 1..N — номера пар (П1, П2, ...). '
    'Жил всего = физическое число жил (для парного: пар × 2). '
    'Идентификаторы должны точно совпадать для кабелей одного сечения (алгоритм объединяет одинаковые барабаны).',
    14)

h2 = ['Марка кабеля', 'Тип', 'Жил всего\n(физ.)',
      'Ид. 1', 'Ид. 2', 'Ид. 3', 'Ид. 4', 'Ид. 5',
      'Ид. 6', 'Ид. 7', 'Ид. 8', 'Ид. 9', 'Ид. 10', 'Примечание']
w2 = [48, 12, 10, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 30]
for i, (h, w) in enumerate(zip(h2, w2), 1):
    set_hdr(ws2, 3, i, h)
    col(ws2, i, w)

COLOR_MAP = {
    'ж/з': YELLOW, 'синий': LIGHT_BLUE, 'чёрный': GRAY_CELL,
    'коричневый': BROWN, 'натуральный': CREAM,
}

comp = [
    # Силовые 5ж
    ('п.10 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1', 'Силовой', 5,  'ж/з','синий','чёрный','коричневый','натуральный','','','','','', 'PE=ж/з, N=синий'),
    ('п.30 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1', 'Силовой', 5,  'ж/з','синий','чёрный','коричневый','натуральный','','','','','', ''),
    ('п.31 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1', 'Силовой', 5,  'ж/з','синий','чёрный','коричневый','натуральный','','','','','', ''),
    ('п.37 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1', 'Силовой', 5,  'ж/з','синий','чёрный','коричневый','натуральный','','','','','', ''),
    ('п.42 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1', 'Силовой', 5,  'ж/з','синий','чёрный','коричневый','натуральный','','','','','', ''),
    ('п.51 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1', 'Силовой', 5,  'ж/з','синий','чёрный','коричневый','натуральный','','','','','', ''),
    ('п.53 Вз-ВВГнг(А)-LS 5х16мк(N,PE)-1',  'Силовой', 5,  'ж/з','синий','чёрный','коричневый','натуральный','','','','','', ''),
    ('п.56 Вз-ВБШвнг(А)-LS 5х16мк(N,PE)-1', 'Силовой', 5,  'ж/з','синий','чёрный','коричневый','натуральный','','','','','', ''),
    ('п.88 Вз-ВВГнг(А)-LS 5х16мк(N,PE)-1',  'Силовой', 5,  'ж/з','синий','чёрный','коричневый','натуральный','','','','','', ''),
    # Силовые 4ж
    ('п.14/1 Вз-ВБШвнг(PE)-LS 4х16мк(N)-1', 'Силовой', 4,  'ж/з','чёрный','коричневый','натуральный','','','','','','', 'PE=ж/з'),
    ('п.14/2 Вз-ВБШвнг(PE)-LS 4х16мк(N)-1', 'Силовой', 4,  'ж/з','чёрный','коричневый','натуральный','','','','','','', ''),
    ('п.14/3 Вз-ВБШвнг(PE)-LS 4х16мк(N)-1', 'Силовой', 4,  'ж/з','чёрный','коричневый','натуральный','','','','','','', ''),
    ('п.14/4 Вз-ВБШвнг(PE)-LS 4х16мк(N)-1', 'Силовой', 4,  'ж/з','чёрный','коричневый','натуральный','','','','','','', ''),
    # Парные
    ('п.20 ТПП 4х2х1',   'Парный', 8,  'П1','П2','П3','П4','',  '',  '',  '',  '', '', '4 пары × 2 жилы = 8 физических жил'),
    ('п.21 ТПП 8х2х0.5', 'Парный', 16, 'П1','П2','П3','П4','П5','П6','П7','П8','', '', '8 пар × 2 жилы = 16 физических жил'),
]

row = 4
for rec in comp:
    is_pair = rec[1] == 'Парный'
    row_bg = PAIR_CLR if is_pair else None
    for ci, v in enumerate(rec, 1):
        bg = row_bg
        if ci >= 4 and v:
            if is_pair:
                bg = PAIR_ID
            else:
                bg = COLOR_MAP.get(v, 'FFFFFF')
        val = v if v != '' else None
        set_data(ws2, row, ci, val, bg=bg,
                 align='center' if ci in (2, 3) else 'left')
    row += 1

ws2.freeze_panes = 'D4'


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: БАРАБАНЫ ТПЖ (входящие)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('3. Барабаны ТПЖ')

title_row(ws3, 1, '🛢  БАРАБАНЫ ТПЖ — исходное сырьё (входящие барабаны с проводом)', 4)
note_row(ws3, 2,
    '⚠  Один ряд = один физический барабан.  '
    'ID — уникальное название (Барабан-1, Б-1250-4 и т.п.).  '
    'Длина — остаток провода в метрах (целое число).  '
    'Сечение и Примечание — справочно, алгоритм не использует.',
    4)

h3 = ['ID барабана', 'Длина, м', 'Сечение (справ.)', 'Примечание']
w3 = [22, 14, 20, 38]
for i, (h, w) in enumerate(zip(h3, w3), 1):
    set_hdr(ws3, 3, i, h)
    col(ws3, i, w)

drums = [
    ('Барабан-1', 10000, '16 мм², мк', ''),
    ('Барабан-2', 10000, '16 мм², мк', ''),
    ('Барабан-3', 10000, '16 мм², мк', ''),
    ('Барабан-4', 11000, '16 мм², мк', ''),
    ('Барабан-5',  9931, '16 мм², мк', 'остаток барабана'),
]
row = 4
for d in drums:
    for ci, v in enumerate(d, 1):
        set_data(ws3, row, ci, v if v != '' else None,
                 align='center' if ci == 2 else 'left')
    row += 1

ws3.freeze_panes = 'A4'


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: ПАРАМЕТРЫ
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('4. Параметры')

title_row(ws4, 1, '⚙️  ПАРАМЕТРЫ АЛГОРИТМА ИЗОЛИРОВАНИЯ И СКРУТКИ', 3)

col(ws4, 1, 46)
col(ws4, 2, 20)
col(ws4, 3, 54)

set_hdr(ws4, 2, 1, 'Параметр', bg=GRAY_HDR, align='left')
set_hdr(ws4, 2, 2, 'Значение', bg=GRAY_HDR)
set_hdr(ws4, 2, 3, 'Примечание / допустимые значения', bg=GRAY_HDR, align='left')

def section(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill(PARAM_HDR)
    cell.font = Font(bold=True, size=10, color='1F4E79')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 18
    return row + 1

def param(ws, row, name, val, comment):
    c1 = ws.cell(row=row, column=1, value=name)
    c1.font = Font(size=10)
    c1.border = thin
    c1.alignment = Alignment(horizontal='left', vertical='center')

    c2 = ws.cell(row=row, column=2, value=val)
    c2.font = Font(bold=True, size=11)
    c2.fill = fill('FFFDE7')
    c2.border = thin
    c2.alignment = Alignment(horizontal='center', vertical='center')

    c3 = ws.cell(row=row, column=3, value=comment)
    c3.font = Font(italic=True, size=9, color='555555')
    c3.border = thin
    c3.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 22
    return row + 1

r = 3
r = section(ws4, r, 'А. ИЗОЛИРОВАНИЕ')
r = param(ws4, r, 'Ёмкость приёмного барабана (изолировка), м', 4500,
          'Максимальная длина жилы на одном мотке после изолирования. '
          '0 = не формировать наборы скрутки (отдать отдельно по каждому цвету).')
r = param(ws4, r, 'Минимальная длина прогона (MIN_SEGMENT), м', 350,
          'Минимально допустимая длина одного прогона изолирования. Физическое ограничение машины.')
r = param(ws4, r, 'Максимум барабанов ТПЖ на одну жилу (MAX_SPLITS)', 1,
          '1 = без стыков — жила одним куском (рекомендуется). '
          '2 и более = допустим стык (сплайс). Обычно стыки запрещены по ТУ.')
r = param(ws4, r, 'Потери на заправку изолировочной линии, м', 5,
          'Дополнительный расход ТПЖ на каждый старт изолировочной машины (заправка нити).')

r = section(ws4, r, 'Б. СКРУТКА')
r = param(ws4, r, 'Ёмкость приёмного барабана (скрутка), м', 2100,
          'Максимальная длина кабеля одного набора скрутки. '
          'Обычно = ёмкости барабана изолировки (чтобы катушки жил совпадали с набором).')

r = section(ws4, r, 'В. ОПТИМИЗАЦИЯ')
r = param(ws4, r, 'Вес минимизации отходов (WASTE_WEIGHT)', 1000,
          'Приоритет минимизации отходов ТПЖ. '
          'Чем выше — тем больше алгоритм заполняет барабаны. Рекомендуется 1000.')
r = param(ws4, r, 'Штраф за смену подающего барабана', 0,
          'Добавить стоимость смены барабана ТПЖ в целевую функцию. '
          '0 = игнорировать. Увеличьте до 10–50 чтобы группировать кабели по барабану.')
r = param(ws4, r, 'Штраф за смену цвета/пары', 0,
          'Добавить стоимость перехода между цветами (переналадка головки). '
          '0 = игнорировать. Увеличьте до 10–100 чтобы минимизировать переходы.')
r = param(ws4, r, 'Лимит времени решения, сек', 60,
          'Ограничение времени работы OR-Tools CP-SAT. '
          'Для 10–15 кабелей хватает 10–30 сек. Для 50+ кабелей увеличьте до 120–300.')

r = section(ws4, r, 'Г. ПАРНЫЕ КАБЕЛИ')
r = param(ws4, r, 'Тип скрутки пар', 'Последовательная',
          'Последовательная = попарная скрутка (П1, П2, ...), затем кабельная скрутка. '
          'Алгоритм автоматически определяет тип по полю «Тип» в листе «2. Состав кабелей».')
r = param(ws4, r, 'Число жил для единицы изолировки (парный кабель)', 2,
          'В паре — 2 жилы одного сечения. Обе изолируются одинаково. '
          'Расход ТПЖ = длина пары × 2. Алгоритм рассчитывает из «Жил всего / 2» наборов пар.')


wb.save('input_v5.xlsx')
print('Saved: input_v5.xlsx')
