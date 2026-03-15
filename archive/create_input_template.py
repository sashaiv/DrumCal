"""
Создаёт шаблон входного Excel-файла для Length Helper.
Запускать один раз для генерации файла input.xlsx
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ─── Цвета оформления ────────────────────────────────────────────────
C_SHEET_TITLE   = '1F4E79'  # тёмно-синий — заголовок листа
C_TABLE_HEADER  = '2E75B6'  # синий       — шапка таблицы
C_SECTION_TITLE = 'BDD7EE'  # светло-синий — заголовок раздела
C_ROW_ODD       = 'FFFFFF'  # белый
C_ROW_EVEN      = 'DEEAF1'  # очень светло-синий
C_WARN          = 'FFE699'  # жёлтый — подсказки/примечания
C_OK            = 'E2EFDA'  # зелёный — "заполнено"

# ─── Вспомогательные функции ─────────────────────────────────────────
def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11):
    return Font(name='Calibri', bold=True, color='FFFFFF', size=size)

def title_font(size=13):
    return Font(name='Calibri', bold=True, color=C_SHEET_TITLE, size=size)

def note_font():
    return Font(name='Calibri', italic=True, color='595959', size=9)

def cell_font(bold=False, size=10):
    return Font(name='Calibri', bold=bold, size=size)

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def center(cell):
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def left(cell):
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_header_row(ws, row, cols, texts, widths=None):
    """Заполнить строку-шапку таблицы."""
    for i, (col, text) in enumerate(zip(cols, texts)):
        c = ws.cell(row=row, column=col, value=text)
        c.font = header_font()
        c.fill = fill(C_TABLE_HEADER)
        c.border = thin_border()
        center(c)
        if widths:
            ws.column_dimensions[get_column_letter(col)].width = widths[i]

def set_title(ws, row, col, text, col_span=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = title_font()
    c.fill = fill(C_SECTION_TITLE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    if col_span:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + col_span - 1
        )

def set_note(ws, row, col, text, col_span=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = note_font()
    c.fill = fill(C_WARN)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if col_span:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + col_span - 1
        )

def data_row(ws, row, col, values, even=False):
    bg = C_ROW_EVEN if even else C_ROW_ODD
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col + i, value=val)
        c.font = cell_font()
        c.fill = fill(bg)
        c.border = thin_border()
        left(c)

def freeze(ws, cell='A3'):
    ws.freeze_panes = cell


# ═══════════════════════════════════════════════════════════════════════
# Лист 1: Заказы
# ═══════════════════════════════════════════════════════════════════════
def sheet_orders(wb):
    ws = wb.active
    ws.title = '1. Заказы'
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 36

    set_title(ws, 1, 1, '📋  ЗАКАЗЫ НА КАБЕЛЬ', col_span=3)
    set_note(ws, 1, 4,
        '⚠  Кабельный журнал — отрезки через запятую (пример: 2000, 1500, 300). '
        'Если журнала нет — оставьте ячейку пустой, длина берётся из столбца Б.',
        col_span=3)

    cols  = [1, 2, 3]
    texts = ['Марка кабеля', 'Длина заказа, м', 'Кабельный журнал (отрезки через запятую)']
    widths = [40, 18, 55]
    set_header_row(ws, 2, cols, texts, widths)

    orders = [
        # Марки должны ТОЧНО совпадать с марками в листе «Состав кабелей»
        ('ВБШвнг(А)-LS 3х2,5ок(N,PE)-1',   6000, '2000, 2000, 300, 300, 300, 150, 600, 350'),
        ('ВБШвнг(А)-FRLS 5х2,5мк(N,PE)-1', 4900, '150, 170, 180, 900, 2000, 1500'),
        ('ВВГнг(А)-LS 4х25мк(N)',            3500, '1000, 800, 700, 500, 500'),
        ('ВВГнг(А)-FRHF 3х25мк(N,PE)',       1200, ''),  # без журнала
        ('ВВГнг(А)-LS 5х25мкг(N,PE)',        2800, '1400, 1400'),
    ]

    for i, (name, length, journal) in enumerate(orders):
        even = (i % 2 == 1)
        data_row(ws, 3 + i, 1, [name, length, journal], even=even)

    freeze(ws, 'A3')


# ═══════════════════════════════════════════════════════════════════════
# Лист 2: Состав кабелей
# ═══════════════════════════════════════════════════════════════════════
def sheet_composition(wb):
    ws = wb.create_sheet('2. Состав кабелей')
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 40

    set_title(ws, 1, 1, '🎨  СОСТАВ КАБЕЛЕЙ — цвета жил', col_span=10)
    set_note(ws, 1, 11,
        '⚠  Сечение — только число (например, 2,5 или 25). '
        'Индекс ТПЖ: ок / мк / мкг. '
        'Огнестойкость: FR или — . '
        'Материал изоляции: LS / HF / PVC и т.п. '
        'Цвета жил должны точно совпадать с названиями в листе П-Ф.',
        col_span=4)

    cols  = [1,  2,            3,         4,              5,                  6,       7,       8,       9,       10,      11]
    texts = ['Марка кабеля', 'Сечение жил', 'Индекс ТПЖ', 'Огнестойкость', 'Материал изоляции',
             'Жила 1', 'Жила 2', 'Жила 3', 'Жила 4', 'Жила 5', 'Примечание']
    widths = [40, 13, 12, 14, 17, 16, 16, 16, 16, 16, 30]
    set_header_row(ws, 2, cols, texts, widths)

    # (марка, сечение, индекс, огнестойкость, материал, ж1, ж2, ж3, ж4, ж5, примечание)
    compositions = [
        # Марки должны ТОЧНО совпадать с марками в листе «Заказы»
        ('ВБШвнг(А)-LS 3х2,5ок(N,PE)-1',    '2,5', 'ок',  '—',  'LS',
         'Натуральная', 'Синяя', 'Желто-зеленая', '', '', 'N=Синяя, PE=Натуральная'),
        ('ВБШвнг(А)-FRLS 5х2,5мк(N,PE)-1',  '2,5', 'мк',  'FR', 'LS',
         'Натуральная', 'Синяя', 'Желто-зеленая', 'Черная', 'Коричневая', 'N=Синяя, PE=Натуральная'),
        ('ВВГнг(А)-LS 4х25мк(N)',             '25',  'мк',  '—',  'LS',
         'Натуральная', 'Синяя', 'Черная', 'Коричневая', '', 'N=Синяя'),
        ('ВВГнг(А)-FRHF 3х25мк(N,PE)',        '25',  'мк',  'FR', 'HF',
         'Натуральная', 'Синяя', 'Желто-зеленая', '', '', 'N=Синяя, PE=Желто-зеленая'),
        ('ВВГнг(А)-LS 5х25мкг(N,PE)',         '25',  'мкг', '—',  'LS',
         'Натуральная', 'Синяя', 'Желто-зеленая', 'Черная', 'Коричневая', 'N=Синяя, PE=Желто-зеленая'),
    ]

    for i, row_data in enumerate(compositions):
        even = (i % 2 == 1)
        data_row(ws, 3 + i, 1, list(row_data), even=even)

    freeze(ws, 'A3')


# ═══════════════════════════════════════════════════════════════════════
# Лист 3: П/Ф (полуфабрикаты — склад материалов)
# Колонки: № | Тип | Сечение, мм² | Индекс | Огнестойкость |
#           Материал изол. | Цвет / Марка кабеля | Длина, м | Примечание (ID)
# ═══════════════════════════════════════════════════════════════════════
def sheet_pf(wb):
    ws = wb.create_sheet('3. П-Ф (склад)')
    N_COLS = 9
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 40

    set_title(ws, 1, 1, '📦  СКЛАД ПОЛУФАБРИКАТОВ И ГОТОВОЙ ПРОДУКЦИИ', col_span=N_COLS)
    set_note(ws, 1, N_COLS + 1,
        '⚠  Тип: ТПЖ / Изолированная / Кабель. '
        'Для ТПЖ: сечение (мм²) + индекс (ок/мк/мкг). '
        'Для Изолированной: + огнестойкость (FR или —) + материал (LS/HF) + цвет. '
        'Для Кабель: марку в колонке «Цвет / Марка кабеля». '
        'Каждая строка = один барабан / бухта / отрезок.',
        col_span=4)

    cols  = list(range(1, N_COLS + 1))
    texts = ['№', 'Тип', 'Сечение, мм²', 'Индекс', 'Огнестойкость',
             'Материал изол.', 'Цвет / Марка кабеля', 'Длина, м', 'Примечание (ID)']
    widths = [5, 15, 14, 10, 14, 15, 26, 10, 20]
    set_header_row(ws, 2, cols, texts, widths)
    ws.row_dimensions[2].height = 36

    # (№, тип, сечение, индекс, огнестойкость, материал, цвет_или_марка, длина, id)
    inventory = [
        # ── ТПЖ 2,5ок ── 6 барабанов
        (1,  'ТПЖ', '2,5', 'ок', '—', '—', '—',  9000, 'Барабан-А1'),
        (2,  'ТПЖ', '2,5', 'ок', '—', '—', '—',  9000, 'Барабан-А2'),
        (3,  'ТПЖ', '2,5', 'ок', '—', '—', '—',  9000, 'Барабан-А3'),
        (4,  'ТПЖ', '2,5', 'ок', '—', '—', '—',  2000, 'Барабан-А4'),
        (5,  'ТПЖ', '2,5', 'ок', '—', '—', '—',  1000, 'Барабан-А5'),
        (6,  'ТПЖ', '2,5', 'ок', '—', '—', '—',   400, 'Барабан-А6'),
        # ── ТПЖ 2,5мк ── 3 барабана (для FR-кабеля ВБШвнг(А)-FRLS, 5×4900м=24500м)
        (7,  'ТПЖ', '2,5', 'мк', '—', '—', '—',  9000, 'Барабан-АМ1'),
        (8,  'ТПЖ', '2,5', 'мк', '—', '—', '—',  9000, 'Барабан-АМ2'),
        (9,  'ТПЖ', '2,5', 'мк', '—', '—', '—',  9000, 'Барабан-АМ3'),
        # ── ТПЖ 25мк ── 3 барабана
        (10, 'ТПЖ', '25',  'мк', '—', '—', '—',  8000, 'Барабан-Б1'),
        (11, 'ТПЖ', '25',  'мк', '—', '—', '—',  8000, 'Барабан-Б2'),
        (12, 'ТПЖ', '25',  'мк', '—', '—', '—',  4000, 'Барабан-Б3'),
        # ── ТПЖ 25мкг ── 2 барабана (для ВВГнг(А)-LS 5х25мкг)
        (13, 'ТПЖ', '25',  'мкг', '—', '—', '—', 8000, 'Барабан-ВГ1'),
        (14, 'ТПЖ', '25',  'мкг', '—', '—', '—', 8000, 'Барабан-ВГ2'),
        # ── Изолированные жилы 2,5ок LS ──
        (15, 'Изолированная', '2,5', 'ок', '—',  'LS', 'Синяя',          6000, ''),
        (16, 'Изолированная', '2,5', 'ок', '—',  'LS', 'Желто-зеленая', 12000, ''),
        (17, 'Изолированная', '2,5', 'ок', '—',  'LS', 'Коричневая',      500, ''),
        # ── Изолированные жилы 25мк LS ──
        (18, 'Изолированная', '25',  'мк', '—',  'LS', 'Серая',          2000, ''),
        # ── Готовый кабель на складе (марки = марки из листа «Заказы») ──
        (19, 'Кабель', '—', '—', '—', '—', 'ВВГнг(А)-FRHF 3х25мк(N,PE)', 300, ''),
        (20, 'Кабель', '—', '—', '—', '—', 'ВВГнг(А)-LS 4х25мк(N)',       500, ''),
    ]

    type_colors = {
        'ТПЖ':           'FFF2CC',   # жёлтый
        'Изолированная': 'E2EFDA',   # зелёный
        'Кабель':        'FCE4D6',   # персиковый
    }

    for i, row_data in enumerate(inventory):
        even = (i % 2 == 1)
        data_row(ws, 3 + i, 1, list(row_data), even=even)
        row_type = row_data[1]
        if row_type in type_colors:
            bg = type_colors[row_type]
            for col in range(1, N_COLS + 1):
                ws.cell(row=3 + i, column=col).fill = fill(bg)

    # Легенда типов
    legend_row = 3 + len(inventory) + 1
    ws.row_dimensions[legend_row].height = 14
    for col, (t, color, label) in enumerate([
        ('ТПЖ',           'FFF2CC', 'ТПЖ — голая жила'),
        ('Изолированная', 'E2EFDA', 'Изолированная — цветная жила'),
        ('Кабель',        'FCE4D6', 'Кабель — готовый кабель'),
    ], start=1):
        c = ws.cell(row=legend_row, column=col, value=label)
        c.font = note_font()
        c.fill = fill(color)
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center', vertical='center')

    freeze(ws, 'A3')


# ═══════════════════════════════════════════════════════════════════════
# Лист 4: Ёмкость барабанов — две секции: Жилы и Кабель
# ═══════════════════════════════════════════════════════════════════════
def _section_header(ws, row, col, text, col_span, note):
    """Заголовок секции внутри листа."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    c.fill = fill('375623')   # тёмно-зелёный
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + col_span - 1)
    set_note(ws, row + 1, col, note, col_span=col_span)
    ws.row_dimensions[row + 1].height = 42


def sheet_drums(wb):
    ws = wb.create_sheet('4. Барабаны')

    # ── Общий заголовок листа ──────────────────────────────────────────
    set_title(ws, 1, 1,
        '🛢  ЁМКОСТЬ БАРАБАНОВ (м) — заголовки столбцов можно менять под свои типы',
        col_span=7)
    ws.row_dimensions[1].height = 20

    # ══════════════════════════════════════════════════════════════════
    # СЕКЦИЯ A: Жилы изолированные (мотки/катушки после изолирования)
    # Ёмкость определяет макс. длину одного прохода изолирования.
    # ══════════════════════════════════════════════════════════════════
    SEC_A_START = 3
    _section_header(
        ws, SEC_A_START, 1,
        '  А. ЖИЛЫ — мотки/катушки после изолирования',
        col_span=7,
        note=(
            '⚠  Заголовки столбцов (Б-400, Б-630, Б-1000) — названия ваших катушек/барабанов. '
            'Измените их на нужные. Добавьте столбцы если нужно больше типов. '
            'Ёмкость = максимальная длина жилы на этом мотке. '
            'Алгоритм выбирает наименьший подходящий моток под каждый прогон изолирования.'
        )
    )

    # Шапка секции A (строка SEC_A_START + 2)
    # Два ключевых столбца: Сечение, мм² | Индекс — затем барабаны
    row_a_hdr = SEC_A_START + 2
    ws.row_dimensions[row_a_hdr].height = 32
    cols_a  = [1, 2, 3, 4, 5]
    texts_a = ['Сечение, мм²', 'Индекс', 'Б-400, м', 'Б-630, м', 'Б-1000, м']
    widths_a = [16, 10, 14, 14, 14]
    set_header_row(ws, row_a_hdr, cols_a, texts_a, widths_a)
    # Подсказка
    c = ws.cell(row=row_a_hdr, column=6,
                value='← Добавляйте столбцы для других типов барабанов')
    c.font = note_font()
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.column_dimensions['F'].width = 38

    # Данные секции A: (сечение, индекс, cap1, cap2, cap3)
    cores_data = [
        ('2,5', 'ок',  2500,  4500,  9000),
        ('2,5', 'мк',  2500,  4500,  9000),
        ('25',  'мк',  1000,  2000,  4000),
        ('25',  'мкг', 1000,  2000,  4000),
    ]
    for i, row_data in enumerate(cores_data):
        data_row(ws, row_a_hdr + 1 + i, 1, list(row_data), even=(i % 2 == 1))

    # ══════════════════════════════════════════════════════════════════
    # СЕКЦИЯ B: Готовый кабель (выходные барабаны)
    # Ёмкость определяет упаковку отрезков кабельного журнала.
    # ══════════════════════════════════════════════════════════════════
    SEC_B_START = row_a_hdr + 1 + len(cores_data) + 2   # пустая строка между секциями
    _section_header(
        ws, SEC_B_START, 1,
        '  Б. КАБЕЛЬ — выходные барабаны для готовой продукции',
        col_span=7,
        note=(
            '⚠  Заголовки столбцов (№10, №12, №14) — типы ваших выходных барабанов. '
            'Измените или добавьте под свои. '
            'Ёмкость = максимальная длина кабеля на барабане данного типа. '
            'Алгоритм пакует отрезки кабельного журнала на наименьшие подходящие барабаны.'
        )
    )

    row_b_hdr = SEC_B_START + 2
    ws.row_dimensions[row_b_hdr].height = 32
    cols_b  = [1, 2, 3, 4]
    texts_b = ['Марка кабеля', '№10, м', '№12, м', '№14, м']
    widths_b = [40, 10, 10, 10]
    # ширину колонки A уже задали выше — обновим
    ws.column_dimensions['A'].width = 40
    set_header_row(ws, row_b_hdr, cols_b, texts_b, widths_b)
    c = ws.cell(row=row_b_hdr, column=5,
                value='← Добавляйте столбцы для других типов барабанов')
    c.font = note_font()
    c.alignment = Alignment(horizontal='left', vertical='center')

    cables_data = [
        # Марки должны ТОЧНО совпадать с марками в листе «Заказы»
        ('ВБШвнг(А)-LS 3х2,5ок(N,PE)-1',   600,  800, 1200),
        ('ВБШвнг(А)-FRLS 5х2,5мк(N,PE)-1', 450,  650,  980),
        ('ВВГнг(А)-LS 4х25мк(N)',            300,  500,  800),
        ('ВВГнг(А)-FRHF 3х25мк(N,PE)',       350,  580,  950),
        ('ВВГнг(А)-LS 5х25мкг(N,PE)',        300,  500,  800),
    ]
    for i, row_data in enumerate(cables_data):
        data_row(ws, row_b_hdr + 1 + i, 1, list(row_data), even=(i % 2 == 1))

    freeze(ws, 'B4')


# ═══════════════════════════════════════════════════════════════════════
# Лист 5: Параметры процесса
# ═══════════════════════════════════════════════════════════════════════
def sheet_params(wb):
    ws = wb.create_sheet('5. Параметры')
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 36

    set_title(ws, 1, 1, '⚙️  ПАРАМЕТРЫ ПРОИЗВОДСТВЕННОГО ПРОЦЕССА', col_span=3)

    cols  = [1, 2, 3]
    texts = ['Параметр', 'Значение', 'Примечание']
    widths = [45, 14, 50]
    set_header_row(ws, 2, cols, texts, widths)

    params = [
        ('Макс. намотка на изолировании, м',      4500,
            'Один проход изолировочной машины — не более этого значения'),
        ('Макс. длина партии скрутки, м',          2100,
            'Один проход скрутки — не более этого значения'),
        ('Минимальная строительная длина, м',        450,
            'Применяется если кабельный журнал не задан'),
        ('Спайка жил разрешена',                   'Нет',
            'Нет — жила должна быть одним куском, спайка запрещена'),
        ('Несколько отрезков на приёмном барабане', 'Да',
            'Да — несколько отрезков кабельного журнала одной партии скрутки '
            'можно принимать на один барабан (отрезки разделены). '
            'Нет — каждый отрезок на отдельный барабан.'),
        ('Стратегия оптимизации',            'Экономия',
            'Экономия = меньше барабанов и отходов; Скорость = меньше переходов по цвету'),
        ('Потери на заправку изолировочной линии, м', 5,
            'На каждый прогон изолирования: ТПЖ снимается на это значение больше плановой длины'),
        ('Запас на обрезку торцов, м',              3,
            'На каждый отрезок журнала в партии скрутки: добавляется к расходу ТПЖ'),
        ('Сохранять порядок кабельного журнала',    'Нет',
            'Нет = алгоритм FFD (сортировка по убыванию); Да = порядок как в журнале'),
        ('Порог предупреждения об остатке→отход, м', 50,
            'Остаток катушки/барабана жилы после использования < порога → предупреждение'),
    ]

    for i, row_data in enumerate(params):
        even = (i % 2 == 1)
        data_row(ws, 3 + i, 1, list(row_data), even=even)

    freeze(ws, 'A3')


# ═══════════════════════════════════════════════════════════════════════
# Главный запуск
# ═══════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    sheet_orders(wb)
    sheet_composition(wb)
    sheet_pf(wb)
    sheet_drums(wb)
    sheet_params(wb)

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'input.xlsx')
    wb.save(out)
    print(f'Файл создан: {out}')

if __name__ == '__main__':
    main()
