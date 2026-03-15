"""
Парсер входного Excel-файла (input_v5.xlsx).

Формат v5 — 4 листа:
  1. Заказы          — марка, длина, гибкий (ДА/НЕТ), кабельный журнал
  2. Состав кабелей  — марка, тип (Силовой/Парный), жил всего, идентификаторы жил/пар
  3. Барабаны ТПЖ   — ID барабана, длина, сечение (справочно)
  4. Параметры       — ключ=значение для ProcessParams

Отличия от root/parser.py (старый формат):
  - Лист «Заказы» добавляет колонку «Гибкий (ДА/НЕТ)» и сдвигает «Журнал» на колонку D
  - Лист «Состав кабелей» упрощён: Тип + Жил всего + Ид.1..10 (без Сечение/Материал)
  - Лист «3. Барабаны ТПЖ» — только ТПЖ (нет склада жил и кабеля)
  - Параметры добавляют CP-SAT поля: min_segment, max_splits, waste_weight

Ограничение v5: нет cross_section/wire_type/fire_resistant/insulation_material
в листе Состав. Для полного конвейера (склад жил, склад кабеля, выходные барабаны)
нужно будет расширить формат в v6.

Совместимость: parse_input_v5() возвращает те же типы, что ожидает planner.plan().
"""
from __future__ import annotations
import re
from typing import List, Tuple, Dict, Optional
import openpyxl

from models import (
    RawWire, InsulatedCore, CableStock, CableOrder,
    DrumType, CoreDrumCapacity, CableDrumCapacity, ProcessParams,
)


# ─── Вспомогательные функции ─────────────────────────────────────────

def _str(val) -> str:
    if val is None:
        return ''
    return str(val).strip()


def _float(val, default: float = 0.0) -> float:
    if val is None:
        return default
    try:
        return float(str(val).replace(',', '.').replace(' ', ''))
    except (ValueError, TypeError):
        return default


def _bool_yes(val) -> bool:
    """'ДА', 'YES', '1', 'TRUE' → True; всё остальное → False."""
    return _str(val).upper().strip() in ('ДА', 'YES', '1', 'TRUE')


def _parse_journal(raw) -> List[float]:
    """
    '2000, 1500, 300' → [2000.0, 1500.0, 300.0]
    Пустая строка / None → []
    """
    s = _str(raw)
    if not s:
        return []
    tokens = re.split(r'[,;\s]+', s)
    result = []
    for t in tokens:
        t = t.strip()
        if t:
            try:
                result.append(float(t))
            except ValueError:
                pass
    return result


def _cross_section_str(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    if not s:
        return ''
    try:
        f = float(s.replace(',', '.'))
        if f == int(f):
            return str(int(f))
        return f'{f:.4g}'.replace('.', ',')
    except (ValueError, TypeError):
        return s


# ─── Листы v5 формата ────────────────────────────────────────────────

def _parse_orders_v5(ws) -> List[CableOrder]:
    """
    Лист '1. Заказы':
      Строка 1 — заголовок, строка 2 — примечание, строка 3 — шапка, строки 4+ — данные.
      A=Марка кабеля  B=Длина, м  C=Гибкий(ДА/НЕТ)  D=Кабельный журнал  E=Примечание
    """
    orders = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        mark = _str(row[0])
        if not mark:
            continue
        length   = _float(row[1] if len(row) > 1 else None)
        flexible = _bool_yes(row[2] if len(row) > 2 else None)
        journal  = _parse_journal(row[3] if len(row) > 3 else None)

        orders.append(CableOrder(
            mark=mark,
            total_length=length,
            journal=journal,
            colors=[],        # заполним позже
            cross_section='', # заполним позже
            flexible=flexible,
        ))
    return orders


def _parse_composition_v5(ws) -> Dict[str, Tuple]:
    """
    Лист '2. Состав кабелей':
      Строки с 4. A=Марка, B=Тип, C=Жил_всего, D..M=Ид.1..10, N=Примечание

    Возвращает:
      {марка: (cable_type, n_cores, [идентификаторы])}
    """
    composition: Dict[str, Tuple] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        mark = _str(row[0])
        if not mark:
            continue
        cable_type = _str(row[1] if len(row) > 1 else None)   # 'Силовой' / 'Парный'
        n_cores    = int(_float(row[2] if len(row) > 2 else None, 0))

        colors = []
        for i in range(3, min(13, len(row))):  # D..M = Ид.1..10
            c = _str(row[i])
            if c:
                colors.append(c)

        composition[mark] = (cable_type, n_cores, colors)
    return composition


def _parse_drums_v5(ws) -> List[RawWire]:
    """
    Лист '3. Барабаны ТПЖ':
      Строки с 4. A=ID, B=Длина, C=Сечение(справ.), D=Примечание

    Возвращает список RawWire. cross_section и wire_type заполняются из параметров.
    """
    drums: List[RawWire] = []
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 1):
        drum_id = _str(row[0])
        if not drum_id:
            continue
        length = _float(row[1] if len(row) > 1 else None)
        if length <= 0:
            continue

        drums.append(RawWire(
            id=drum_id,
            name=drum_id,
            cross_section='',   # заполним после парсинга параметров
            wire_type='',
            length=length,
        ))
    return drums


def _parse_params_v5(ws) -> Tuple[ProcessParams, str, str]:
    """
    Лист '4. Параметры':
      Строки с 3. A=Параметр, B=Значение, C=Примечание

    Возвращает (params, cross_section, wire_type) — глобальные для всего расчёта.
    cross_section и wire_type нужны для RawWire, когда в листе Барабаны нет сечения.
    """
    params = ProcessParams()
    cross_section = ''
    wire_type = ''

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        key = _str(row[0]).lower()
        val = row[1] if len(row) > 1 else None

        # v5 параметры изолирования
        if 'ёмкость' in key and ('приём' in key or 'барабан' in key or 'изолир' in key):
            params.max_insulation_run = _float(val, params.max_insulation_run)
        elif 'min_segment' in key or 'минимальн' in key and 'прогон' in key:
            params.min_segment = _float(val, params.min_segment)
        elif 'max_splits' in key or 'максимум барабанов' in key:
            params.max_splits = int(_float(val, params.max_splits))
        elif 'потери' in key and 'заправк' in key:
            params.insulation_startup_loss_m = _float(val, params.insulation_startup_loss_m)
        # v5 параметры скрутки
        elif 'ёмкость' in key and 'скрутк' in key:
            params.max_twisting_run = _float(val, params.max_twisting_run)
        elif 'строительн' in key or ('минимальн' in key and ('длин' in key or 'кабел' in key)):
            params.min_construction_length = _float(val, params.min_construction_length)
        elif 'допуск' in key or ('запас' in key and 'торц' in key):
            params.length_tolerance_m = _float(val, params.length_tolerance_m)
        elif 'порог' in key and 'отход' in key:
            params.waste_warning_threshold_m = _float(val, params.waste_warning_threshold_m)
        elif 'порядок' in key and 'журнал' in key:
            params.keep_journal_order = _bool_yes(val)
        # CP-SAT параметры
        elif 'waste_weight' in key or ('приоритет' in key and 'отход' in key):
            params.waste_weight = int(_float(val, params.waste_weight))
        elif 'time_limit' in key or ('лимит' in key and 'врем' in key):
            params.cpsat_time_limit = _float(val, params.cpsat_time_limit)
        # Глобальные параметры жилы
        elif 'сечение' in key and ('жил' in key or 'тпж' in key or 'провод' in key):
            cross_section = _cross_section_str(val)
        elif 'индекс' in key and ('тпж' in key or 'жил' in key or 'тип' in key):
            wire_type = _str(val)

    return params, cross_section, wire_type


# ─── Главная функция парсинга (v5) ───────────────────────────────────

def parse_input_v5(filepath: str):
    """
    Читает input_v5.xlsx и возвращает все объекты для планирования.

    Возвращает:
        orders, raw_wires, insulated_cores, cable_stock,
        core_drum_caps, cable_drum_caps, params

    Примечания по v5 → v6 эволюции:
      - insulated_cores = [] (склад жил в v5 не поддерживается)
      - cable_stock     = [] (склад готового кабеля в v5 не поддерживается)
      - core_drum_caps  = [] (ёмкости катушек-приёмников в v5 не заданы;
                              берётся max_insulation_run из параметров)
      - cable_drum_caps = [] (ёмкости выходных барабанов в v5 не заданы)

    Для работы полного конвейера (склад жил, склад кабеля, выходные барабаны)
    создайте input_v6.xlsx с дополнительными листами или используйте root/parser.py.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    def _sheet(keyword: str):
        for name in wb.sheetnames:
            if keyword.lower() in name.lower():
                return wb[name]
        raise KeyError(f'Лист с "{keyword}" не найден в {filepath}')

    # Параметры парсим первыми — нужен cross_section / wire_type для RawWire
    params, cross_section, wire_type = _parse_params_v5(_sheet('Параметр'))

    orders     = _parse_orders_v5(_sheet('Заказ'))
    composition = _parse_composition_v5(_sheet('Состав'))
    raw_wires  = _parse_drums_v5(_sheet('Барабан'))

    # Применяем глобальный wire_key ко всем барабанам ТПЖ
    for rw in raw_wires:
        rw.cross_section = cross_section
        rw.wire_type     = wire_type

    # Обогащаем заказы цветами из листа состава
    warnings = []
    for order in orders:
        if order.mark in composition:
            cable_type, n_cores, colors = composition[order.mark]
            order.colors = colors
            order.cross_section = cross_section
            order.wire_type = wire_type
            # В v5 нет fire_resistant/insulation_material — используем дефолты
        else:
            warnings.append(
                f'⚠ Марка "{order.mark}" не найдена в листе "Состав кабелей".'
            )

    if warnings:
        print('\n'.join(warnings))

    # В v5 нет складов и барабанных ёмкостей
    insulated_cores: List[InsulatedCore] = []
    cable_stock: List[CableStock] = []
    core_drum_caps: List[CoreDrumCapacity] = []
    cable_drum_caps: List[CableDrumCapacity] = []

    return (orders, raw_wires, insulated_cores, cable_stock,
            core_drum_caps, cable_drum_caps, params)


# ─── Совместимость: парсер старого формата ───────────────────────────
# (оставлен из root/parser.py для обратной совместимости)

def _normalize_fr(val: str) -> str:
    s = val.strip().upper()
    return 'FR' if s == 'FR' else ''


def _drum_type_name(header: str) -> str:
    return re.sub(r',?\s*м\s*$', '', _str(header)).strip()


def _parse_orders_old(ws) -> List[CableOrder]:
    """Лист '1. Заказы' старого формата: строки с 3."""
    orders = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        mark = _str(row[0])
        if not mark:
            continue
        length = _float(row[1])
        journal = _parse_journal(row[2] if len(row) > 2 else None)
        orders.append(CableOrder(
            mark=mark, total_length=length, journal=journal,
            colors=[], cross_section='',
        ))
    return orders


def _parse_composition_old(ws) -> Dict[str, Tuple]:
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    color_col_indices = []
    if header_row:
        for i in range(len(header_row)):
            if _str(header_row[i]).lower().startswith('жила'):
                color_col_indices.append(i)

    composition: Dict[str, Tuple] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        mark = _str(row[0])
        if not mark:
            continue
        cross_section       = _cross_section_str(row[1] if len(row) > 1 else None)
        wire_type           = _str(row[2] if len(row) > 2 else None)
        fire_resistant      = _normalize_fr(_str(row[3] if len(row) > 3 else None))
        insulation_material = _str(row[4] if len(row) > 4 else None)
        colors = [_str(row[i]) for i in color_col_indices if i < len(row) and _str(row[i])]
        composition[mark] = (cross_section, wire_type, fire_resistant, insulation_material, colors)
    return composition


def _parse_pf_old(ws) -> Tuple[List[RawWire], List[InsulatedCore], List[CableStock]]:
    raw_wires: List[RawWire] = []
    insulated: List[InsulatedCore] = []
    cable_stock: List[CableStock] = []
    counters = {'ТПЖ': 0, 'Изолированная': 0, 'Кабель': 0}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        pf_type             = _str(row[1] if len(row) > 1 else None)
        cross_section       = _cross_section_str(row[2] if len(row) > 2 else None)
        wire_type           = _str(row[3] if len(row) > 3 else None)
        fire_resistant      = _normalize_fr(_str(row[4] if len(row) > 4 else None))
        insulation_material = _str(row[5] if len(row) > 5 else None)
        name_or_mark        = _str(row[6] if len(row) > 6 else None)
        length              = _float(row[7] if len(row) > 7 else None)
        note                = _str(row[8] if len(row) > 8 else None)

        if pf_type not in ('ТПЖ', 'Изолированная', 'Кабель') or length <= 0:
            continue

        counters[pf_type] += 1
        item_id = note if note else f'{pf_type}-{counters[pf_type]:03d}'

        if pf_type == 'ТПЖ':
            raw_wires.append(RawWire(
                id=item_id, name=f'ТПЖ {cross_section}{wire_type}',
                cross_section=cross_section, wire_type=wire_type, length=length,
            ))
        elif pf_type == 'Изолированная':
            color = name_or_mark
            ins_name = f'{color} {cross_section}{wire_type}'
            if fire_resistant:
                ins_name += f' {fire_resistant}'
            ins_name += f' {insulation_material}'
            insulated.append(InsulatedCore(
                id=item_id, name=ins_name, color=color,
                cross_section=cross_section, wire_type=wire_type,
                insulation_material=insulation_material, fire_resistant=fire_resistant,
                length=length,
            ))
        else:
            cable_stock.append(CableStock(id=item_id, cable_mark=name_or_mark, length=length))

    return raw_wires, insulated, cable_stock


def _parse_drums_old(ws) -> Tuple[List[CoreDrumCapacity], List[CableDrumCapacity]]:
    core_caps: List[CoreDrumCapacity] = []
    cable_caps: List[CableDrumCapacity] = []
    rows = list(ws.iter_rows(values_only=True))

    def _find_header_row(keyword: str) -> Optional[int]:
        for i, row in enumerate(rows):
            if row and _str(row[0]) == keyword:
                return i
        return None

    def _read_core_section(header_idx: int):
        header_row = rows[header_idx]
        drum_names = []
        for col_val in header_row[2:]:
            s = _str(col_val)
            if not s or '←' in s or 'Добавляйте' in s:
                break
            drum_names.append(_drum_type_name(s))

        data_rows = []
        for row in rows[header_idx + 1:]:
            if not row:
                break
            cross_raw = _str(row[0])
            if not cross_raw or cross_raw.startswith('←') or not cross_raw[0].isdigit():
                break
            cross = _cross_section_str(row[0])
            wtype = _str(row[1]) if len(row) > 1 else ''
            wire_key = f'{cross}{wtype}'
            capacities = [_float(row[j + 2]) for j in range(len(drum_names))]
            data_rows.append((wire_key, capacities))
        return drum_names, data_rows

    def _read_cable_section(header_idx: int):
        header_row = rows[header_idx]
        drum_names = []
        for col_val in header_row[1:]:
            s = _str(col_val)
            if not s or '←' in s:
                break
            drum_names.append(_drum_type_name(s))

        data_rows = []
        for row in rows[header_idx + 1:]:
            if not row:
                break
            key = _str(row[0])
            if not key:
                break
            capacities = [_float(row[j + 1]) for j in range(len(drum_names))]
            data_rows.append((key, capacities))
        return drum_names, data_rows

    idx_a = _find_header_row('Сечение, мм²')
    if idx_a is not None:
        drum_names, data_rows = _read_core_section(idx_a)
        for wire_key, capacities in data_rows:
            drum_types = sorted(
                [DrumType(dn, cap) for dn, cap in zip(drum_names, capacities) if cap > 0],
                key=lambda d: d.capacity,
            )
            core_caps.append(CoreDrumCapacity(wire_key=wire_key, drum_types=drum_types))

    idx_b = _find_header_row('Марка кабеля')
    if idx_b is not None:
        drum_names, data_rows = _read_cable_section(idx_b)
        for cable_mark, capacities in data_rows:
            drum_types = sorted(
                [DrumType(dn, cap) for dn, cap in zip(drum_names, capacities) if cap > 0],
                key=lambda d: d.capacity,
            )
            cable_caps.append(CableDrumCapacity(cable_mark=cable_mark, drum_types=drum_types))

    return core_caps, cable_caps


def _parse_params_old(ws) -> ProcessParams:
    params = ProcessParams()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        key = _str(row[0]).lower()
        val = row[1] if len(row) > 1 else None

        if 'изолирован' in key:
            params.max_insulation_run = _float(val, params.max_insulation_run)
        elif 'скрутк' in key or 'партии' in key:
            params.max_twisting_run = _float(val, params.max_twisting_run)
        elif 'строительн' in key or 'минимальн' in key:
            params.min_construction_length = _float(val, params.min_construction_length)
        elif 'заправк' in key or 'startup' in key:
            params.insulation_startup_loss_m = _float(val, params.insulation_startup_loss_m)
        elif 'допуск' in key and 'торц' in key:
            params.length_tolerance_m = _float(val, params.length_tolerance_m)
        elif 'порядок' in key and 'журнал' in key:
            params.keep_journal_order = _bool_yes(val)
        elif 'отход' in key or 'порог' in key:
            params.waste_warning_threshold_m = _float(val, params.waste_warning_threshold_m)

    return params


def parse_input(filepath: str):
    """
    Читает input.xlsx старого формата (5 листов).
    Аналог root/parser.py, адаптированный под new/models.py (добавлен flexible).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    def _sheet(keyword: str):
        for name in wb.sheetnames:
            if keyword.lower() in name.lower():
                return wb[name]
        raise KeyError(f'Лист с "{keyword}" не найден в {filepath}')

    orders      = _parse_orders_old(_sheet('Заказы'))
    composition = _parse_composition_old(_sheet('Состав'))
    raw_wires, insulated_cores, cable_stock = _parse_pf_old(_sheet('П-Ф'))
    core_drum_caps, cable_drum_caps = _parse_drums_old(_sheet('Барабан'))
    params = _parse_params_old(_sheet('Параметр'))

    warnings = []
    for order in orders:
        if order.mark in composition:
            cross_section, wire_type, fire_resistant, insulation_material, colors = composition[order.mark]
            order.cross_section = cross_section
            order.wire_type = wire_type
            order.fire_resistant = fire_resistant
            order.insulation_material = insulation_material
            order.colors = colors
        else:
            warnings.append(f'⚠ Марка "{order.mark}" не найдена в "Состав кабелей".')

    if warnings:
        print('\n'.join(warnings))

    return (orders, raw_wires, insulated_cores, cable_stock,
            core_drum_caps, cable_drum_caps, params)
